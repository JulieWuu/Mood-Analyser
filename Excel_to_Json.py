import json

from openpyxl import load_workbook

EXCEL_PATH = "EVERYDAYlog.xlsx"
TEMPLATE = {"mood": 0,
            "date": 0,
            "month": 0,
            "day": 0,
            "week": 0,
            "comment": ''}

def get_next_date(date_pair : tuple[int, int]) -> tuple[int, int]:
    month = date_pair[0]
    date = date_pair[1]

    if (month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10) and date == 31:
        return tuple((month + 1, 1))
    elif month == 2 and date == 28:
        return tuple((3, 1))
    elif (month == 4 or month == 6 or month == 9 or month == 11) and date == 30:
        return tuple((month + 1, 1))
    elif month == 12 and date == 31:
        return tuple((1, 1))
    else:
        return tuple((month, date + 1))

def scrape_cell(cell, data : dict) -> dict:
    data["comment"] = cell.value
    data["day"] = (cell.column + 5) % 7 + 1
    data["week"] = (cell.row - 1) // 2

    colour = cell.font.color

    if colour.type == 'rgb':
        # blue
        if colour.rgb == 'FF0070C0':
            data["mood"] = 1
        # green
        elif colour.rgb == 'FF00B050':
            data["mood"] = 2
        else:
            raise FormatNotRecognised("Cell has a custom colour that is not blue or green")

    elif colour.type == 'theme':
        # yellow
        if colour.theme == 6:
            data["mood"] = 3
        else:
            raise FormatNotRecognised("Cell has a theme colour that is not yellow")

    return data


class FormatNotRecognised(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(message)

    def __str__(self):
        return f"{self.message}"


class ExcelReader:
    def __init__(self, path, template, js_filename):
        self.excel_log = load_workbook(filename=path, read_only=True).active

        self.template = template
        self.scraped_data = {}
        self.js_filename = js_filename


    def scrape_sheet(self, min_row=1, max_row=1000):
        index = 1
        date_pair = tuple((1, 1))

        for row in self.excel_log.iter_rows(min_row=min_row, max_row=max_row):
            for cell in row:
                try:
                    if cell.row <=2 or (cell.row == 3 and cell.column <= 3) or (cell.row >= 107 and cell.column >= 5):
                        continue
                    elif cell.row % 2 == 1:
                        data = self.template.copy()
                        try:
                            data = scrape_cell(cell, data)
                            data["month"] = date_pair[0]
                            data["date"] = date_pair[1]

                        except FormatNotRecognised:
                            print(f"Cannot recognise colour format for number {index} cell.")
                            continue

                        self.scraped_data[index] = data
                        index += 1
                        date_pair = get_next_date(date_pair)
                except AttributeError:
                    continue

        print(f"Total number of records: {index - 1}")


    def to_json(self):
        with open(self.js_filename, "w") as js_file:
            json.dump(self.scraped_data, js_file)


if __name__ == "__main__":
    excel_reader = ExcelReader(EXCEL_PATH, TEMPLATE, "json_mood.json")
    excel_reader.scrape_sheet()
    excel_reader.to_json()
